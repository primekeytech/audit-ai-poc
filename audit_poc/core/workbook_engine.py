# ============================================
# AUDIT POC - Workbook Engine
# ============================================
# This file handles loading and populating
# Benjamin's Excel work program.
#
# CRITICAL RULES:
# 1. NEVER recreate workbook from scratch
# 2. Always load EXISTING template
# 3. Only populate columns F, G, H
# 4. Use data_only=FALSE to preserve formulas
# 5. Save as NEW file - never overwrite template
# ============================================

import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import copy
from datetime import datetime

# ============================================
# MAIN WORKBOOK FUNCTIONS
# ============================================

def load_template(template_path: str):
    """
    Loads the existing work program template.
    
    CRITICAL: data_only=False preserves ALL formulas!
    If we used data_only=True, formulas would be lost!
    
    Args:
        template_path: Path to Benjamin's template.xlsx
        
    Returns:
        openpyxl.Workbook: Loaded workbook object
    """
    
    # Check template exists
    if not Path(template_path).exists():
        raise FileNotFoundError(
            f"Work program template not found at: {template_path}\n"
            f"Please add Benjamin's template.xlsx to the workbook/ folder"
        )
    
    print(f"Loading workbook template: {template_path}")
    
    # CRITICAL: data_only=False preserves formulas!
    # keep_vba=True preserves any Excel macros
    workbook = load_workbook(
        template_path,
        data_only=False,    # FALSE = keep formulas intact!
        keep_vba=False      # No VBA macros expected
    )
    
    print(f"Template loaded successfully!")
    print(f"Sheets found: {workbook.sheetnames}")
    
    return workbook


def populate_workbook(workbook, ai_results: list, config: dict):
    """
    Populates the work program with AI analysis results.
    Only touches columns F, G, H - nothing else!
    
    Args:
        workbook: Loaded openpyxl workbook
        ai_results: List of AI analysis results
            Each result = {
                "row": row_number,
                "score": 1-4,
                "artifact": "document reference",
                "notes": "AI explanation"
            }
        config: Configuration dictionary
        
    Returns:
        openpyxl.Workbook: Populated workbook
    """
    
    # Get column settings from config
    score_col = config["workbook"]["score_column"]      # F
    artifact_col = config["workbook"]["artifact_column"] # G
    notes_col = config["workbook"]["notes_column"]       # H
    
    # Get the first/main sheet
    # TODO: Update sheet name once we see Benjamin's template
    sheet = workbook.active
    
    print(f"Populating workbook - {len(ai_results)} controls to process")
    
    # Loop through each AI result
    for result in ai_results:
        
        # Get the row number for this control
        row_num = result.get("row")
        
        if not row_num:
            print(f"Warning: No row number for result - skipping")
            continue
        
        # Populate Column F - Score (1-4)
        score_cell = f"{score_col}{row_num}"
        sheet[score_cell] = result.get("score", 3)
        
        # Populate Column G - Artifact reference
        artifact_cell = f"{artifact_col}{row_num}"
        sheet[artifact_cell] = result.get("artifact", "")
        
        # Populate Column H - AI Notes
        notes_cell = f"{notes_col}{row_num}"
        sheet[notes_cell] = result.get("notes", "")
        
        print(f"Row {row_num}: Score={result.get('score')} populated")
    
    print("Workbook population complete!")
    return workbook


def save_workbook(workbook, output_path: str):
    """
    Saves the populated workbook as a NEW file.
    NEVER overwrites the original template!
    
    Args:
        workbook: Populated openpyxl workbook
        output_path: Where to save the new file
        
    Returns:
        str: Path where file was saved
    """
    
    # Create outputs directory if it doesn't exist
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Add timestamp to filename to avoid overwriting
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Build output filename with timestamp
    output_file = str(output_path).replace(
        ".xlsx", 
        f"_{timestamp}.xlsx"
    )
    
    # Save the workbook
    workbook.save(output_file)
    
    print(f"Workbook saved to: {output_file}")
    return output_file


def get_control_rows(workbook, config: dict) -> list:
    """
    Reads the work program and returns all control rows
    that need to be analysed.
    
    This function reads the existing workbook to find
    which rows contain audit controls.
    
    Args:
        workbook: Loaded openpyxl workbook
        config: Configuration dictionary
        
    Returns:
        list: List of control dictionaries
            Each = {
                "row": row_number,
                "control_id": "2.9",
                "control_question": "Does policy exist?",
                "document_ref": "2.9"
            }
    """
    
    # Get the active sheet
    sheet = workbook.active
    
    # Store all control rows
    controls = []
    
    # TODO: Update column letters once we see Benjamin's template
    # For now using placeholder column positions
    # These will be updated when template arrives
    
    # Loop through rows starting from row 2 (skip header)
    for row_num in range(2, sheet.max_row + 1):
        
        # Get control question from column C (placeholder)
        control_cell = sheet[f"C{row_num}"]
        
        # Skip empty rows
        if not control_cell.value:
            continue
        
        # Build control dictionary
        control = {
            "row": row_num,
            "control_id": str(sheet[f"A{row_num}"].value or ""),
            "control_question": str(control_cell.value or ""),
            "document_ref": str(sheet[f"B{row_num}"].value or ""),
        }
        
        controls.append(control)
    
    print(f"Found {len(controls)} control rows in workbook")
    return controls


def get_findings(workbook, config: dict) -> list:
    """
    Reads populated workbook and returns all
    rows with score 4 (findings).
    
    These findings will go into the PDF report.
    
    Args:
        workbook: Populated workbook
        config: Configuration dictionary
        
    Returns:
        list: All score 4 rows as finding dictionaries
    """
    
    sheet = workbook.active
    findings = []
    
    # Get score column letter from config
    score_col = config["workbook"]["score_column"]  # F
    finding_threshold = config["workbook"]["finding_threshold"]  # 4
    
    # Loop through all rows
    for row_num in range(2, sheet.max_row + 1):
        
        # Get score from column F
        score_cell = sheet[f"{score_col}{row_num}"]
        score = score_cell.value
        
        # Check if this is a finding (score 4)
        if score == finding_threshold:
            
            # Build finding dictionary
            # TODO: Update column letters to match Benjamin's template
            finding = {
                "row": row_num,
                "control_id": str(sheet[f"A{row_num}"].value or ""),
                "control_question": str(sheet[f"C{row_num}"].value or ""),
                "score": score,
                "artifact": str(sheet[f"G{row_num}"].value or ""),
                "notes": str(sheet[f"H{row_num}"].value or ""),
                # These columns will be mapped once template arrives
                "finding_header": str(sheet[f"L{row_num}"].value or ""),
                "risk_description": str(sheet[f"M{row_num}"].value or ""),
                "recommendation": str(sheet[f"N{row_num}"].value or ""),
            }
            
            findings.append(finding)
    
    print(f"Found {len(findings)} findings (score 4) in workbook")
    return findings
